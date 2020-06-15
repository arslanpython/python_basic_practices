from openpyxl import Workbook, load_workbook

from openpyxl import Workbook, load_workbook


def create_xls_write_headers(filepath):
    wb = Workbook()
    sheet = wb.active

    cols = [
        "Fruit",
        "Flavour",
        "Expiration"
    ]

    for index, value in enumerate(cols):
        sheet.cell(row=1, column=index+1).value = value

    wb.save(filepath)


def write_xls(filepath, dictionary):
    wb = load_workbook(filepath)
    sheet = wb.active

    for i, x in enumerate(dictionary):
        for idx, value in enumerate(x.values()):
            sheet.cell(row=i+2, column=idx+1).value = value

    wb.save(filepath)


things = [
    {
        "Fruit": "Orange",
        "Flavour": "Good",
        "Expiration": "21May20"
    },
    {
        "Fruit": "Apple",
        "Flavour": "Good",
        "Expiration": "19May20"
    },
    {
        "Fruit": "Banana",
        "Flavour": "Regular",
        "Expiration": "16May20"
    }
]


fp = 'test.xlsx'
create_xls_write_headers(fp)
write_xls(fp, things)


# def create_xls_write_headers(filepath):
#     wb = Workbook()
#     sheet = wb.active
#
#     cols = [
#         "Fruit",
#         "Flavour",
#         "Expiration"
#     ]
#
#     for index, value in enumerate(cols):
#         sheet.cell(row=1, column=index + 1).value = value
#
#     wb.save(filepath)
#
#
# sheet_row_index = 0
#
#
# def write_xls(file_path, item):
#     global sheet_row_index
#
#     wb = load_workbook(file_path)
#     sheet = wb.active
#
#     for idx, value in enumerate(item.values()):
#         sheet.cell(row=sheet_row_index + 2, column=idx + 1).value = value
#
#     wb.save(file_path)
#     sheet_row_index += 1
#
#
# things = [
#     {
#         "Fruit": "Orange",
#         "Flavour": "Good",
#         "Expiration": "21May20"
#     },
#     {
#         "Fruit": "Apple",
#         "Flavour": "Good",
#         "Expiration": "19May20"
#     },
#     {
#         "Fruit": "Banana",
#         "Flavour": "Regular",
#         "Expiration": "16May20"
#     }
# ]
#
#
# fp = 'test.xlsx'
# create_xls_write_headers(fp)
# write_xls(fp, dict(things.pop()))
# write_xls(fp, dict(things.pop()))
# write_xls(fp, dict(things.pop()))
# print(sheet_row_index)